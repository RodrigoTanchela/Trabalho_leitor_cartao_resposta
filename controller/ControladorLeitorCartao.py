import openpyxl


from ProjetoPin3.model.ConfiguracaoQuestao import ConfiguracaoQuestao
from ProjetoPin3.model.ConfiguracaoOpcoes import ConfiguracoesOpcoes
from ProjetoPin3.View.TelaGeracaoCartaoResposta import TelaGeracaoCartaoResposta

import json
import statistics
import math
import cv2
import numpy as np
# from openpyxl import *

class ControladorLeitorCartao:
    def __init__(self):
        self.visaoLeitorCartaoResposta = None
        self.aluno = None
        self.cartaoResposta = None
        self.configuracaoOpcoes = None
        self.configuracaoQuestao = None
        self.geracaoPlanilha = None
        self.questoes = None
        self.keypoints = None

    def abrir_explorador_de_arquivos(self, arquivo_selecionado):
        self.visaoLeitorCartaoResposta.atualizar_label(arquivo_selecionado)

    def actionPreview(self):
        img = self.visaoLeitorCartaoResposta.getImage()
        keypoints, img_cinza = self.identificandoPontos(img)
        image = self.construcaoFormularioRetangular(keypoints, img_cinza)
        self.visaoLeitorCartaoResposta.previewConfiguracoes(image)

    def importar_cartao_resposta(self,  arquivo_selecionado):
        self.telaGeracaoCartaoResposta.atualizar_label(arquivo_selecionado)

    def definirConfiguracao(self, configuracao):
        numeroPerguntas = configuracao['numeroPerguntas']
        numeroOpcoes = configuracao['numeroOpcoes']
        margemLateral = configuracao['margemLateral']
        margemSuperior = configuracao['margemSuperior']
        larguraMarcador = configuracao['larguraMarcador']
        alturaMarcador = configuracao['alturaMarcador']
        espacamentoPerguntas = configuracao['espacamentoPerguntas']
        espacamentoResposta = configuracao['espacamentoResposta']

        self.keypoints = configuracao['keypoints']
        keypoints = configuracao['keypoints']
        keypoint = keypoints[2]
        posicaoHorizontal = int(keypoint.pt[0])
        posicaoVertical = int(keypoint.pt[1])

        self.configuracaoQuestao = ConfiguracaoQuestao(espacamentoPerguntas, numeroPerguntas, margemSuperior, margemLateral, posicaoHorizontal, posicaoVertical)
        self.configuracaoOpcoes = ConfiguracoesOpcoes(espacamentoResposta, numeroOpcoes, larguraMarcador, alturaMarcador)


    def actionImportacao(self):
        self.telaGeracaoCartaoResposta = TelaGeracaoCartaoResposta(self)
        self.telaGeracaoCartaoResposta.iniciar()

    def tamahoAreaRecorte(self):
        largura = int(self.keypoints[3].pt[0]) - int(self.keypoints[2].pt[0])
        altura = int(self.keypoints[0].pt[1]) - int(self.keypoints[2].pt[1])
        return largura, altura

    def leituraRespostas(self, image, img_cinza, caminho_arquivo):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value = 'Alunos')
        for aluno in range(len(image)):
            sheet.cell(row=aluno+2, column = 1, value=f'Aluno{aluno+1}')
            respostas_alunos = self.identificar_respostas(image[aluno], aluno)
            for questao in range (self.configuracaoQuestao.get_numero()):
                sheet.cell(row=1, column= questao+2, value=f'Questão {questao+1}')
                respostas = str(respostas_alunos[questao]).strip('[]')
                sheet.cell(row=aluno+2, column=questao+2, value = respostas)
        self.remover_caractere(workbook)
        workbook.save(caminho_arquivo)
        self.visaoLeitorCartaoResposta.popupInformativo(f'Dados salvos em {caminho_arquivo}')

    def remover_caractere(self, workbook):
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = cell.value.replace("'", '')
    def salvandoDadosTela(self, dados, nome_arquivo):
        with open(nome_arquivo, 'w') as arquivo:
            json.dump(dados, arquivo, indent=4)

    def identificar_respostas(self, image, aluno):
        respostas = []
        respostas_Questao = []
        keypoints, img = self.identificandoPontos(image)
        if(len(keypoints) != 4):
            configuracoes = self.getConfiguracaoPadrao()
            x = int(self.keypoints[2].pt[0])
            y = int(self.keypoints[2].pt[1])
            self.telaGeracaoCartaoResposta.popupInformativo(f'Ouve dificuldade em indentificar os marcadores da prova do aluno: {aluno+1}'
                                                            f' sera gerado normalmente porém deve ser verificado com mais atenção posteriormente')
        else:
            configuracoes = self.configuracoesFolhaResposta(keypoints)
            x = int(keypoints[2].pt[0])
            y = int(keypoints[2].pt[1])

        retangulo_width = configuracoes['largura']
        retangulo_height = configuracoes['altura']

        distancia_horizontal = configuracoes['espacamentoPerguntas']
        distancia_vertical = configuracoes['espacamentoRespostas']

        start_x = int(x + configuracoes['margemLateral'])
        start_y = int(y + configuracoes['margemSuperior'])

        for i in range(self.configuracaoQuestao.get_numero()):
            for j in range(self.configuracaoOpcoes.get_numero()):
                x1, y1 = start_x, start_y
                x2, y2 = start_x + retangulo_width, start_y + retangulo_height
                retangulo = image[y1:y2, x1:x2]

                media_canais = cv2.mean(retangulo)
                valorCanais = statistics.mean(media_canais)
                if valorCanais <= 145:
                    numero_retangulo = j+1
                    letra = self.numero_para_letra(numero_retangulo)
                    respostas_Questao.append(letra)
                elif valorCanais > 145 and valorCanais < 150:
                    letra = self.telaGeracaoCartaoResposta.popupCaixaTesto(f'Houve dificuldade em identificar opção correta {j+1} na questão {i+1} da prova do aluno {aluno+1}')
                    respostas_Questao.append(letra)
                    break;


                cv2.rectangle(image, (x1, y1), (x2, y2), (0, 255, 0), 1)

                start_x += retangulo_width + distancia_horizontal


            respostas.append(respostas_Questao)
            respostas_Questao = []
            start_x = int(x + configuracoes['margemLateral'])
            start_y += retangulo_height + distancia_vertical
        return respostas


    def verificar_cor(area, cor):
        area_bgr = cv2.cvtColor(area, cv2.COLOR_BGR2RGB)
        cor_bgr = np.array(cor, dtype=np.uint8)
        return np.any(np.all(area_bgr == cor_bgr, axis=-1))

    def identificandoPontos(self, imagem):
        params = cv2.SimpleBlobDetector_Params()
        params.filterByColor = True
        params.blobColor = 0
        params.filterByArea = True
        params.minRepeatability = 16
        params.maxArea = 350

        img_cinza = cv2.cvtColor(np.array(imagem), cv2.COLOR_BGR2GRAY)

        detector = cv2.SimpleBlobDetector_create(params)
        keypoints = detector.detect(img_cinza)
        keypoints = self.ordenarKeyPoints(keypoints)

        return keypoints, img_cinza

    def construcaoFormularioRetangular(self, keypoints, image):
        keypoint = keypoints[2]
        x, y = int(keypoint.pt[0]), int(keypoint.pt[1])
        largura, altura = self.visaoLeitorCartaoResposta.getLarguraMarcador(), self.visaoLeitorCartaoResposta.getAlturaMarcador()
        espacamento_resposta, espacamento_pergunta = self.visaoLeitorCartaoResposta.getEspacamentoResposta() , self.visaoLeitorCartaoResposta.getEspacamentoPergunta()
        x_inicio, y_inicio = int(x + self.visaoLeitorCartaoResposta.getMargemLateral()), int(y + self.visaoLeitorCartaoResposta.getMargemSuperior())
        color = (0, 255, 0)
        thickness = 1
        x = x_inicio
        y = y_inicio
        for _ in range(self.visaoLeitorCartaoResposta.getNumeroPerguntas()):
            for _ in range(self.visaoLeitorCartaoResposta.getNumeroOpcoes()):
                x2, y2 = x + largura, y + altura
                cv2.rectangle(image, (x, y), (x2, y2), color, thickness)
                x += largura + espacamento_resposta
            x = x_inicio
            y += altura + espacamento_pergunta

        return image

    def configuracoesFolhaResposta(self, keypoints):
        escala = self.calculaEscala(keypoints)

        espacamentoRespostas = self.calcularedimensaoNecessaria(self.configuracaoOpcoes.get_espacamento(), escala['comprimentoConfiguracao'], escala['comprimentoRespostas'])
        largura = self.calcularedimensaoNecessaria(self.configuracaoOpcoes.get_largura(), escala['comprimentoConfiguracao'], escala['comprimentoRespostas'])
        margemLateral = self.calcularedimensaoNecessaria(self.configuracaoQuestao.get_margemLateral(), escala['comprimentoConfiguracao'], escala['comprimentoRespostas'])

        espacamentoPerguntas = self.calcularedimensaoNecessaria(self.configuracaoOpcoes.get_espacamento(), escala['alturaCofiguracao'], escala['alturaRespostas'])
        altura = self.calcularedimensaoNecessaria(self.configuracaoOpcoes.get_altura(), escala['alturaCofiguracao'], escala['alturaRespostas'])
        margemSuperior = self.calcularedimensaoNecessaria(self.configuracaoQuestao.get_margemSuperior(), escala['alturaCofiguracao'], escala['alturaRespostas'])

        return {
            'espacamentoRespostas': int(round(espacamentoRespostas)),
            'largura': int(round(largura)),
            'margemLateral': int(math.ceil(margemLateral)),
            'espacamentoPerguntas': int(round(espacamentoPerguntas)),
            'altura': int(round(altura)),
            'margemSuperior': int(math.ceil(margemSuperior))
        }

    def calcularedimensaoNecessaria(self, valorConfiguracao, keypointConfiguracao, keypointResposta):
        return (keypointResposta * valorConfiguracao) / keypointConfiguracao

    def calculaEscala(self, keypoints):
        comprimentoConfiguracao, alturaCofiguracao = self.calculaDimensaoKeypoins(self.keypoints)
        comprimentoRespostas, alturaRespostas = self.calculaDimensaoKeypoins(keypoints)
        comprimento = alturaCofiguracao - alturaRespostas
        altura = comprimentoConfiguracao - comprimentoRespostas
        return {
            'comprimentoConfiguracao': comprimentoConfiguracao,
            'alturaCofiguracao': alturaCofiguracao,
            'comprimentoRespostas': comprimentoRespostas,
            'alturaRespostas': alturaRespostas,
            'comprimento': comprimento,
            'altura': altura

        }

    def calculaDimensaoKeypoins(self, keypoints):
        comprimento = keypoints[0].pt[0] - keypoints[2].pt[0]
        altura = keypoints[2].pt[1] - keypoints[3].pt[1]
        return comprimento, altura

    def numero_para_letra(self, numero):
        if 1 <= numero <= 26:
            return chr(ord("a") + numero - 1)
        else:
            return '?'

    def ordenarKeyPoints(self, keypoints):
        keypoints_ordenados = sorted(keypoints, key=lambda keypoint: (keypoint.pt[0], keypoint.pt[1]), reverse=True)
        return keypoints_ordenados

    def rotacionar_imagem(self, imagem, angulo):
        altura, largura = imagem.shape[:2]
        ponto_central = (largura // 2, altura // 2)
        matriz_rotacao = cv2.getRotationMatrix2D(ponto_central, angulo, 1.0)
        imagem_rotacionada = cv2.warpAffine(imagem, matriz_rotacao, (largura, altura), flags=cv2.INTER_LINEAR)

        return imagem_rotacionada

    def calculaAnguloInterno(self, keypoints):
        xfictio, yficticio = keypoints[2].pt[0], keypoints[3].pt[1]
        y = yficticio - keypoints[2].pt[1]
        x = xfictio - keypoints[3].pt[0]
        h = math.sqrt(((y ** 2) + (x ** 2)))
        cos_theta = y / h

        theta_radianos = math.acos(cos_theta)
        sinal = x - keypoints[3].pt[0]
        theta_graus = math.degrees(theta_radianos)
        graus = round(theta_graus)
        if(sinal >= 0 and graus <= 10):
            return -graus
        elif(sinal < 0 and graus <= 10):
            return +graus
        return 0

    def getConfiguracaoPadrao(self):
        espacamentoRespostas = self.configuracaoOpcoes.get_espacamento()
        largura = self.configuracaoOpcoes.get_largura()
        margemLateral = self.configuracaoQuestao.get_margemLateral()

        espacamentoPerguntas = self.configuracaoOpcoes.get_espacamento()
        altura = self.configuracaoOpcoes.get_altura()
        margemSuperior = self.configuracaoQuestao.get_margemSuperior()

        return {
            'espacamentoRespostas': int(round(espacamentoRespostas)),
            'largura': int(round(largura)),
            'margemLateral': int(math.ceil(margemLateral)),
            'espacamentoPerguntas': int(round(espacamentoPerguntas)),
            'altura': int(round(altura)),
            'margemSuperior': int(math.ceil(margemSuperior))
        }

